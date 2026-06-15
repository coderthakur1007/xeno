import tempfile
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from app.agents.copilot import MarketingCopilotGraph


PROOF_PROMPTS = [
    "Convert new shoppers into second-purchase customers.",
    "Bring back customers who have not purchased in 90 days.",
    "Find high value shoppers and launch an exclusive drop.",
    "Win back platinum skincare buyers in Mumbai with WhatsApp.",
    "Target email customers who clicked but have not converted.",
    "Prevent churn among loyal customers inactive for 60 days.",
    "Promote footwear sale to gold tier shoppers in Bangalore.",
    "Recover customers with failed SMS deliveries.",
    "Grow average order value for grocery buyers with a bundle offer.",
    "Invite recent first-time women shoppers to a second purchase campaign.",
    "Launch an RCS campaign for engaged apparel customers.",
    "Reward campaign buyers who converted after the last promotion.",
]


class CopilotProofService:
    def __init__(self, db: Session):
        self.db = db

    def run_prompt_matrix(self, tenant_id: uuid.UUID) -> list[dict[str, Any]]:
        graph = MarketingCopilotGraph(self.db)
        results = []
        for prompt in PROOF_PROMPTS:
            plan = graph.plan(tenant_id, prompt)
            results.append(
                {
                    "prompt": prompt,
                    "audience_size": plan["audience_size"],
                    "detected_intents": ", ".join(plan["segment"]["filters"].get("detected_intents", [])),
                    "rules": plan["segment"]["filters"].get("rules", []),
                    "channels": ", ".join(plan.get("recommended_channels", plan["strategy"].get("channels", []))),
                    "conversion_probability": plan.get("conversion_probability", plan["strategy"].get("conversion_probability", 0)),
                    "expected_conversions": plan.get("expected_conversions", plan["strategy"].get("expected_conversions", 0)),
                    "sql_text": plan["segment"]["sql_text"],
                    "explainability": " | ".join(plan["explainability"]),
                    "customer_insights": plan.get("customer_insights", {}).get("summary", ""),
                    "analytics_insights": "; ".join(plan.get("analytics_insights", [])),
                    "execution_ready": plan.get("execution_readiness", {}).get("ready", False),
                    "agents_run": len(plan.get("agents", [])),
                }
            )
        return results

    def build_excel(self, tenant_id: uuid.UUID) -> Path:
        rows = self.run_prompt_matrix(tenant_id)
        workbook = Workbook()
        summary = workbook.active
        summary.title = "Proof Summary"
        prompt_sheet = workbook.create_sheet("Prompt Coverage")
        rules_sheet = workbook.create_sheet("Safe SQL Rules")
        training_sheet = workbook.create_sheet("Training Signals")
        agent_sheet = workbook.create_sheet("Agent Performance")

        self._write_summary(summary, rows)
        self._write_prompt_coverage(prompt_sheet, rows)
        self._write_rules(rules_sheet, rows)
        self._write_training_signals(training_sheet)
        self._write_agent_performance(agent_sheet, rows)
        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A2"
            for column in sheet.columns:
                max_len = max(len(str(cell.value or "")) for cell in column[:80])
                sheet.column_dimensions[column[0].column_letter].width = min(max(max_len + 2, 12), 48)
        out_dir = Path(tempfile.gettempdir()) / "xeno-proof-reports"
        out_dir.mkdir(parents=True, exist_ok=True)
        output = out_dir / f"xeno_copilot_proof_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
        workbook.save(output)
        return output

    def _style_header(self, sheet) -> None:
        fill = PatternFill("solid", fgColor="0F766E")
        for cell in sheet[1]:
            cell.fill = fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical="center")

    def _write_summary(self, sheet, rows: list[dict[str, Any]]) -> None:
        sheet.append(["Metric", "Value"])
        sheet.append(["Prompts tested", len(rows)])
        sheet.append(["Prompts with generated SQL", sum(1 for row in rows if "SELECT c.id" in row["sql_text"])])
        sheet.append(["Average conversion probability", f"=AVERAGE('Prompt Coverage'!F2:F{len(rows)+1})"])
        sheet.append(["Expected conversions", f"=SUM('Prompt Coverage'!G2:G{len(rows)+1})"])
        sheet.append(["Execution-ready prompts", sum(1 for row in rows if row.get("execution_ready"))])
        sheet.append(["Agents per prompt", rows[0].get("agents_run", 7) if rows else 7])
        sheet.append(["Report generated UTC", datetime.now(timezone.utc).isoformat()])
        self._style_header(sheet)
        sheet["A10"] = "Channel Forecast"
        sheet["A10"].font = Font(bold=True, size=14)
        sheet.append([])
        sheet.append([])
        sheet.append(["Prompt", "Expected conversions"])
        for row in rows:
            sheet.append([row["prompt"][:44], row["expected_conversions"]])
        chart = BarChart()
        chart.title = "Expected conversions by prompt"
        chart.y_axis.title = "Conversions"
        chart.x_axis.title = "Prompt"
        data = Reference(sheet, min_col=2, min_row=12, max_row=12 + len(rows))
        cats = Reference(sheet, min_col=1, min_row=13, max_row=12 + len(rows))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 7
        chart.width = 18
        sheet.add_chart(chart, "D10")

    def _write_prompt_coverage(self, sheet, rows: list[dict[str, Any]]) -> None:
        sheet.append([
            "Prompt", "Detected intents", "Audience size", "Channels", "Rules",
            "Conversion probability", "Expected conversions", "Explainability",
            "Customer Insights", "Execution Ready",
        ])
        for row in rows:
            sheet.append(
                [
                    row["prompt"],
                    row["detected_intents"],
                    row["audience_size"],
                    row["channels"],
                    str(row["rules"]),
                    row["conversion_probability"],
                    row["expected_conversions"],
                    row["explainability"],
                    row.get("customer_insights", ""),
                    "Yes" if row.get("execution_ready") else "No",
                ]
            )
        self._style_header(sheet)
        for cell in sheet["F"][1:]:
            cell.number_format = "0.0%"

    def _write_rules(self, sheet, rows: list[dict[str, Any]]) -> None:
        sheet.append(["Prompt", "Rule field", "Operator", "Value", "Generated parameterized SQL"])
        for row in rows:
            for rule in row["rules"]:
                sheet.append([row["prompt"], rule["field"], rule["operator"], rule["value"], row["sql_text"]])
        self._style_header(sheet)

    def _write_training_signals(self, sheet) -> None:
        sheet.append(["Feature", "Source table", "Used for", "Example meaning"])
        rows = [
            ["recency_days", "orders", "segmentation, churn, conversion prediction", "Days since most recent paid order"],
            ["frequency", "orders", "RFM, repeat-purchase targeting", "Number of paid orders"],
            ["monetary_value", "orders", "LTV and VIP targeting", "Total paid order amount"],
            ["channel_affinity", "communication_events", "channel optimization", "Open/read/click/conversion rates by channel"],
            ["category_affinity", "customers.attributes", "content personalization", "Preferred category from profile and purchases"],
            ["loyalty_tier", "customers.attributes", "offer strategy", "Bronze, silver, gold, or platinum membership"],
            ["delivery_failure_rate", "communication_events", "deliverability recovery", "Provider failure history"],
            ["campaign_conversion_label", "communication_events", "model training", "Whether a message produced conversion event"],
        ]
        for row in rows:
            sheet.append(row)
        self._style_header(sheet)

    def _write_agent_performance(self, sheet, rows: list[dict[str, Any]]) -> None:
        """New sheet showing the 7-agent pipeline run details."""
        sheet.append(["Prompt", "Agents Run", "Execution Ready", "Analytics Insights"])
        for row in rows:
            sheet.append([
                row["prompt"],
                row.get("agents_run", 7),
                "Yes" if row.get("execution_ready") else "No",
                row.get("analytics_insights", "")[:200],
            ])
        self._style_header(sheet)
